"""Generate the multilingual Bhagavad Gita Quarto website.

Every language gets its own subtree of the project:

    en/index.qmd  en/glossary.qmd  en/chapters/chapter-01.qmd ...
    it/index.qmd  it/glossary.qmd  it/chapters/chapter-01.qmd ...

plus a small root index.qmd that sends the visitor to one of them.
A single `quarto render` still builds the whole site.

To add or change a language, edit LANGS below.  Everything else
(_quarto.yml, sidebars, navbar, flags, page titles) follows from it.
"""

from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict
import html
import json
import re
import csv
import shutil

ROOT = Path(__file__).resolve().parents[1]
SRC_XLSX = ROOT / 'data' / 'bhagavadgita_ai_refined.xlsx'
OUT = ROOT
FLAG_SRC_DIR = ROOT / 'data'
FLAG_OUT_DIR = ROOT / 'assets' / 'flags'

GENERATED_FILE_NOTE = '''<!--
GENERATED FILE. DO NOT EDIT DIRECTLY.

Edit instead:
- data/glossary.en.csv / data/glossary.it.csv for glossary content
- data/bhagavadgita_ai_refined.xlsx for verses/translations
- tools/0_generate_site.py for generation logic and language settings
- styles.css for visual style

Then run:
    run.bat
-->
'''

SITE_TITLE = 'Bhagavad Gita'
SITE_TITLE_SANSKRIT = 'भगवद्गीता'

# Conventional romanized chapter titles, aligned with the Sanskrit titles in the
# spreadsheet.  Some editions use close variants, for example Karma-Sannyasa Yoga
# for Chapter 5 or Dhyana Yoga for Chapter 6.  They are Sanskrit, so they are
# shared by every language.
CHAPTER_TITLES_ROMAN = {
    1: "Arjuna Vishada Yoga",
    2: "Sankhya Yoga",
    3: "Karma Yoga",
    4: "Jnana-Karma-Sannyasa Yoga",
    5: "Sannyasa Yoga",
    6: "Atma-Samyama Yoga",
    7: "Jnana-Vijnana Yoga",
    8: "Akshara-Brahma Yoga",
    9: "Raja-Vidya Raja-Guhya Yoga",
    10: "Vibhuti Yoga",
    11: "Vishvarupa-Darshana Yoga",
    12: "Bhakti Yoga",
    13: "Kshetra-Kshetrajna Vibhaga Yoga",
    14: "Guna-Traya Vibhaga Yoga",
    15: "Purushottama Yoga",
    16: "Daivasura-Sampad Vibhaga Yoga",
    17: "Shraddha-Traya Vibhaga Yoga",
    18: "Moksha-Sannyasa Yoga",
}

# ---------------------------------------------------------------------------
# Language configuration.  This is the only place that needs editing to change
# the wording of a language, to point it at a different glossary, or to add a
# third language.
#
# Translation state is detected, not declared: a verse with an empty cell in its
# language column falls back to English, and the notice at the top of a page
# reflects what that page actually contains.  Translating is therefore only ever
# a matter of filling in data/bhagavadgita_ai_refined.xlsx (column
# ai_refined_it, next to ai_refined_en) and data/glossary.it.csv.  The notices
# disappear on their own once nothing falls back.
# ---------------------------------------------------------------------------

LANGS = {
    'en': {
        'label': 'English',
        'flag': 'en.png',
        'glossary_csv': ROOT / 'data' / 'glossary.en.csv',
        'glossary_groups': ['Concepts', 'Social and ritual terms', 'Names and epithets'],
        'ui': {
            'home': 'Home',
            'chapters': 'Chapters',
            'glossary': 'Glossary',
            'chapter': 'Chapter',
            'verse': 'verse',
            'verses': 'verses',
            'previous_chapter': 'Previous chapter',
            'next_chapter': 'Next chapter',
            'sanskrit': 'Sanskrit',
            'usage': 'Lightly marked Sanskrit terms show a short gloss on hover or tap.',
            'download': 'Download the English PDF',
            'pdf_filename': 'bhagavad-gita-en.pdf',
            'glossary_forms': 'Forms in the text:',
            # Notices shown while a language is incomplete; unused for the
            # language the text is written in.
            'pending_all': '',
            'pending_some': '',
            'pending_glossary': '',
        },
        # Longer descriptive prose, kept as content rather than interface text.
        'site_note': (
            'The Bhagavad Gita is part of the Mahabharata, Book 6, Bhishma Parva, '
            'chapters 23-40. This English draft was prepared from Google Translate '
            'output, revised with GPT-5.4-mini, GPT-5.5, Claude-Opus-4.8, plus some '
            'manual revision.'
        ),
        'glossary_note': (
            'A compact guide to transliterated Sanskrit terms and recurring names '
            'used in the English rendering.'
        ),
    },
    'it': {
        'label': 'Italiano',
        'flag': 'it.png',
        'glossary_csv': ROOT / 'data' / 'glossary.it.csv',
        'glossary_groups': ['Concetti', 'Termini sociali e rituali', 'Nomi ed epiteti'],
        'ui': {
            'home': 'Home',
            'chapters': 'Capitoli',
            'glossary': 'Glossario',
            'chapter': 'Capitolo',
            'verse': 'versetto',
            'verses': 'versetti',
            'previous_chapter': 'Capitolo precedente',
            'next_chapter': 'Capitolo successivo',
            'sanskrit': 'Sanscrito',
            'usage': 'I termini sanscriti evidenziati mostrano una breve glossa al passaggio del mouse o al tocco.',
            'download': 'Scarica il PDF in italiano',
            'pdf_filename': 'bhagavad-gita-it.pdf',
            'glossary_forms': 'Forme nel testo:',
            'pending_all': (
                'La traduzione italiana non è ancora disponibile: '
                'il testo mostrato è quello inglese.'
            ),
            'pending_some': (
                'Traduzione italiana in corso: i versetti non ancora tradotti '
                'sono mostrati in inglese.'
            ),
            'pending_glossary': (
                'Il glossario non è ancora tradotto: le voci sono mostrate in inglese.'
            ),
        },
        # TODO: the wording is a literal translation of the English note and
        # still describes how the English draft was made; update the provenance
        # to describe the Italian rendering.
        'site_note': (
            'La Bhagavad Gita fa parte del Mahabharata, Libro 6, Bhishma Parva, '
            'capitoli 23-40. Questa bozza inglese è stata preparata a partire '
            "dall'output di Google Translate, rivista con GPT-5.4-mini, GPT-5.5, "
            'Claude-Opus-4.8, oltre a qualche revisione manuale.'
        ),
        'glossary_note': (
            'Una guida essenziale ai termini sanscriti traslitterati e ai nomi '
            'ricorrenti nella traduzione italiana.'
        ),
    },
}

LANG_ORDER = ['en', 'it']
DEFAULT_LANG = 'en'

# Spreadsheet column holding the translation for each language.  A language
# whose column is absent (or whose cell is empty) falls back to English.
VERSE_COLUMNS = {
    'en': ['ai_refined_en'],
    'it': ['ai_refined_it', 'italian_translation', 'it_translation', 'traduzione_italiana'],
}

SPEAKER_MAP = {
    'धृतराष्ट्र': '(Dhritarashtra)',
    'सञ्जय': '(Sanjaya)',
    'अर्जुन': '(Arjuna)',
    'श्रीभगवान्': '(Krishna)',
}

LETTER = "A-Za-z0-9ĀāĪīŪūṚṛṜṝḶḷṂṃḤḥÑñṄṅṆṇṬṭḌḍŚśṢṣÇçḸḹẎẏ"


def as_text(v):
    return '' if v is None else str(v).strip()


def esc(v):
    return html.escape(as_text(v), quote=True)


def speaker_display(v):
    text = as_text(v)
    return SPEAKER_MAP.get(text, text)


# ---------------------------------------------------------------------------
# Glossary: one instance per language, each with its own matching regex.
# ---------------------------------------------------------------------------


class Glossary:
    def __init__(self, path):
        self.path = path
        self.entries = self._load(path)
        self.regex, self.lookup = self._build_regex()

    @staticmethod
    def _load(path):
        entries = []
        if not path.exists():
            raise FileNotFoundError(
                f'Missing glossary file: {path}. '
                'Create it (a copy of data/glossary.en.csv is a fine starting point).'
            )
        with open(path, encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            required = {'id', 'term', 'group', 'definition', 'variants'}
            missing = required.difference(reader.fieldnames or [])
            if missing:
                raise ValueError(f'Missing columns in {path.name}: {sorted(missing)}')
            for row in reader:
                entries.append({
                    'id': row['id'].strip(),
                    'term': row['term'].strip(),
                    'group': row['group'].strip(),
                    'definition': row['definition'].strip(),
                    'variants': [
                        v.strip()
                        for v in row['variants'].split('|')
                        if v.strip()
                    ],
                })
        return entries

    def _build_regex(self):
        variant_map = []
        for item in self.entries:
            for variant in item['variants']:
                variant_map.append((variant, item))
        variant_map.sort(key=lambda x: len(x[0]), reverse=True)
        pattern = '|'.join(re.escape(v) for v, _ in variant_map)
        regex = re.compile(rf'(?<![{LETTER}])({pattern})(?![{LETTER}])', flags=re.IGNORECASE)
        lookup = {v.casefold(): item for v, item in variant_map}
        return regex, lookup

    def payload(self):
        payload = []
        for item in self.entries:
            clean = dict(item)
            clean['variants'] = sorted(set(clean['variants']))
            payload.append(clean)
        return payload

    def forms_used(self, texts):
        """Which declared variants really occur in `texts`, keyed by entry id.

        The variant lists are deliberately a superset: they carry diacritic
        spellings and compounds that a given translation may never use, so that
        editing a verse cannot silently break a highlight.  The glossary page,
        though, promises "forms in the text", so it must show only what the
        reader will actually meet there.
        """
        seen = defaultdict(set)
        for text in texts:
            for match in self.regex.finditer(as_text(text)):
                item = self.lookup.get(match.group(0).casefold())
                if item is not None:
                    seen[item['id']].add(match.group(0).casefold())
        return {
            item['id']: [
                v for v in item['variants'] if v.casefold() in seen.get(item['id'], ())
            ]
            for item in self.entries
        }

    def annotate(self, text):
        text = as_text(text)
        if not text:
            return ''
        out = []
        pos = 0
        for match in self.regex.finditer(text):
            out.append(html.escape(text[pos:match.start()], quote=True))
            token = match.group(0)
            item = self.lookup.get(token.casefold())
            if item is None:
                out.append(html.escape(token, quote=True))
            else:
                # The highlighted text already identifies the term.  Keep only the
                # definition in the tooltip, so it is not repeated on small screens.
                gloss = item['definition']
                out.append(
                    '<span class="glossary-term" tabindex="0" '
                    f'data-gloss="{html.escape(gloss, quote=True)}" '
                    f'aria-label="{html.escape(gloss, quote=True)}" '
                    f'title="{html.escape(gloss, quote=True)}" '
                    f'data-glossary-id="{html.escape(item["id"], quote=True)}">'
                    f'{html.escape(token, quote=True)}</span>'
                )
            pos = match.end()
        out.append(html.escape(text[pos:], quote=True))
        return ''.join(out)


def html_lines(v, glossary=None):
    if glossary is not None:
        return '<br>\n'.join(glossary.annotate(line) for line in as_text(v).splitlines())
    return '<br>\n'.join(esc(v).splitlines())


# ---------------------------------------------------------------------------
# Source data
# ---------------------------------------------------------------------------


def read_verses(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['Sheet1']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    required = ['chapter', 'chapter_title_sanskrit', 'verse', 'reference', 'speaker', 'sanskrit_sloka', 'ai_refined_en']
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f'Missing required columns: {missing}')

    # Resolve one spreadsheet column per language (None when not present yet).
    lang_cols = {}
    for lang in LANG_ORDER:
        lang_cols[lang] = next((c for c in VERSE_COLUMNS.get(lang, []) if c in idx), None)

    verses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch = row[idx['chapter']]
        verse = row[idx['verse']]
        if ch is None or verse is None:
            continue
        item = {
            'chapter': int(ch),
            'chapter_title_sanskrit': as_text(row[idx['chapter_title_sanskrit']]),
            'verse': int(verse),
            'reference': f"{int(ch)}.{int(verse)}",
            'speaker': speaker_display(row[idx['speaker']]),
            'sanskrit_sloka': as_text(row[idx['sanskrit_sloka']]),
        }
        for lang, col in lang_cols.items():
            item[f'translation_{lang}'] = as_text(row[idx[col]]) if col else ''
        verses.append(item)
    return verses, lang_cols


def translation_for(verse, lang):
    """Return (text, fell_back) for a verse in `lang`.

    An untranslated verse shows the English text instead of nothing, so a
    partially translated language stays readable throughout.
    """
    text = verse.get(f'translation_{lang}')
    if text or lang == DEFAULT_LANG:
        return text, False
    return verse.get(f'translation_{DEFAULT_LANG}', ''), True


# ---------------------------------------------------------------------------
# Shared snippets
# ---------------------------------------------------------------------------


def language_span(lang):
    """Label for a language link, tagged so that styles.css can put the flag in
    front of it and lang-switch.html can recognise the link.

    The flag is a CSS background rather than an <img> on purpose: paths inside
    styles.css resolve against the site root, so one rule is correct for pages
    at any depth, in local previews and under a GitHub Pages project subpath
    alike.
    """
    return f'[{LANGS[lang]["label"]}]{{.gita-lang data-lang="{lang}"}}'


def language_links_yaml(indent):
    """Sidebar entries offering each language."""
    pad = ' ' * indent
    lines = []
    for lang in LANG_ORDER:
        lines.append(f"{pad}- text: '{language_span(lang)}'")
        lines.append(f'{pad}  href: {lang}/index.qmd')
    return '\n'.join(lines)


def pending_banner(lang, verses):
    """Notice describing how much of `verses` is still English.

    Nothing is emitted once the whole set is translated, so the notices clear
    themselves chapter by chapter as the spreadsheet is filled in.
    """
    if lang == DEFAULT_LANG:
        return []
    pending = sum(1 for v in verses if translation_for(v, lang)[1])
    if not pending:
        return []
    key = 'pending_all' if pending == len(verses) else 'pending_some'
    note = LANGS[lang]['ui'].get(key, '')
    if not note:
        return []
    return [f'<div class="lang-pending">{esc(note)}</div>', '']


def glossary_banner(lang, glossaries):
    """Notice for a glossary that is still a verbatim copy of the English one."""
    if lang == DEFAULT_LANG:
        return []
    base = {e['id']: e['definition'] for e in glossaries[DEFAULT_LANG].entries}
    translated = any(
        e['definition'] != base.get(e['id'], e['definition'])
        for e in glossaries[lang].entries
    )
    note = LANGS[lang]['ui'].get('pending_glossary', '')
    if translated or not note:
        return []
    return [f'<div class="lang-pending">{esc(note)}</div>', '']


def page_header(title, lang, sidebar_id=None):
    parts = ['---', f'title: "{title}"', f'lang: {lang}']
    if sidebar_id:
        parts.append(f'sidebar: {sidebar_id}')
    parts += ['---', '']
    return parts


# ---------------------------------------------------------------------------
# _quarto.yml
# ---------------------------------------------------------------------------


def write_quarto_yml(chapters):
    sorted_ch = sorted(chapters)

    def chapter_entries(lang, indent):
        pad = ' ' * indent
        return '\n'.join(
            f'{pad}- text: "{ch} · {CHAPTER_TITLES_ROMAN.get(ch, f"Chapter {ch}")}"\n'
            f'{pad}  href: {lang}/chapters/chapter-{ch:02d}.qmd'
            for ch in sorted_ch
        )

    nav_lang = DEFAULT_LANG  # relocalized client-side by lang-switch.html
    navbar_flags = '\n'.join(
        f"      - text: '{language_span(lang)}'\n"
        f'        href: {lang}/index.qmd'
        for lang in LANG_ORDER
    )

    sidebars = []
    for lang in LANG_ORDER:
        ui = LANGS[lang]['ui']
        # The chapters sit directly in the sidebar rather than inside a
        # collapsible section: the site is bilingual per URL prefix, so a
        # "<language> translation" wrapper would only hide the chapter list
        # behind an extra click.
        sidebars.append(f'''    - id: {lang}
      style: docked
      contents:
{language_links_yaml(8)}
        - text: "---"
        - text: "{ui['home']}"
          href: {lang}/index.qmd
        - text: "---"
{chapter_entries(lang, 8)}
        - text: "---"
        - text: "{ui['glossary']}"
          href: {lang}/glossary.qmd''')

    quarto_yml = f'''# GENERATED FILE. DO NOT EDIT DIRECTLY.
# Edit tools/0_generate_site.py (see LANGS) and run run.bat.
project:
  type: website
  output-dir: docs
  resources:
    - .nojekyll
    - assets

website:
  title: "{SITE_TITLE}"
  search: false
  page-navigation: true
  navbar:
    left:
      - text: "{LANGS[nav_lang]['ui']['chapters']}"
        menu:
{chapter_entries(nav_lang, 10)}
      - text: "{LANGS[nav_lang]['ui']['glossary']}"
        href: {nav_lang}/glossary.qmd
    right:
{navbar_flags}
  sidebar:
{chr(10).join(sidebars)}

format:
  html:
    theme: cosmo
    css: styles.css
    include-after-body:
      - glossary-tooltip.html
      - lang-config.html
      - lang-switch.html
    toc: false
    smooth-scroll: true
    anchor-sections: false
    link-external-newwindow: true
'''
    (OUT / '_quarto.yml').write_text(quarto_yml, encoding='utf-8')


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------


def write_lang_config(chapters):
    """Emit the settings lang-switch.html needs at runtime.

    Kept separate so the switcher logic itself stays a plain, hand-editable
    file, while the wording and the page list stay single-sourced from LANGS.
    """
    pages = ['index.html', 'glossary.html'] + [
        f'chapters/chapter-{ch:02d}.html' for ch in sorted(chapters)
    ]
    # Navbar labels are written in DEFAULT_LANG; map them to the other languages.
    labels = {}
    for lang in LANG_ORDER:
        if lang == DEFAULT_LANG:
            continue
        labels[lang] = {
            LANGS[DEFAULT_LANG]['ui'][key]: LANGS[lang]['ui'][key]
            for key in ('chapters', 'glossary')
        }
    config = {
        'langs': LANG_ORDER,
        'default': DEFAULT_LANG,
        'labels': labels,
        'pages': pages,
    }
    body = json.dumps(config, ensure_ascii=False, indent=2)
    (OUT / 'lang-config.html').write_text(
        '<!-- GENERATED FILE. Edit LANGS in tools/0_generate_site.py instead. -->\n'
        f'<script>\nwindow.GITA_LANG = {body};\n</script>\n',
        encoding='utf-8',
    )


def write_root_index():
    """Small landing page: sends visitors to their language, and works as a
    plain chooser when JavaScript is off."""
    choices = '\n'.join(
        f'<a class="lang-choice" href="{lang}/index.html">'
        f'<span class="gita-lang" data-lang="{lang}">{esc(LANGS[lang]["label"])}</span></a>'
        for lang in LANG_ORDER
    )
    langs_js = json.dumps(LANG_ORDER)
    parts = [
        '---',
        f'title: "{SITE_TITLE}"',
        'sidebar: false',
        'page-navigation: false',
        '---',
        '',
        GENERATED_FILE_NOTE,
        '',
        f'<div class="chapter-title-sanskrit">{SITE_TITLE_SANSKRIT}</div>',
        '',
        '<div class="lang-chooser">',
        choices,
        '</div>',
        '',
        '<script>',
        '(() => {',
        f'  const langs = {langs_js};',
        f'  const fallback = "{DEFAULT_LANG}";',
        '  const stored = (() => { try { return localStorage.getItem("gita-lang"); } catch (e) { return null; } })();',
        '  const preferred = (navigator.languages || [navigator.language || ""])',
        '    .map(tag => String(tag).slice(0, 2).toLowerCase())',
        '    .find(code => langs.includes(code));',
        '  const target = langs.includes(stored) ? stored : (preferred || fallback);',
        '  const base = location.pathname.replace(/index\\.html?$/, "").replace(/\\/?$/, "/");',
        '  location.replace(base + target + "/index.html" + location.hash);',
        '})();',
        '</script>',
        '',
    ]
    (OUT / 'index.qmd').write_text('\n'.join(parts), encoding='utf-8')


def write_index(lang, chapters):
    cfg = LANGS[lang]
    ui = cfg['ui']
    cards = []
    for ch in sorted(chapters):
        count = len(chapters[ch])
        verse_word = ui['verse'] if count == 1 else ui['verses']
        title_roman = CHAPTER_TITLES_ROMAN.get(ch, f'{ui["chapter"]} {ch}')
        cards.append(
            f'<a class="chapter-card" href="chapters/chapter-{ch:02d}.html">'
            f'<span class="chapter-number">{esc(ui["chapter"])} {ch}</span>'
            f'<span class="chapter-title">{esc(title_roman)}</span>'
            f'<span class="chapter-count">{count} {esc(verse_word)}</span></a>'
        )
    parts = page_header(SITE_TITLE, lang, sidebar_id=lang)
    parts += [
        GENERATED_FILE_NOTE,
        '',
        f'<div class="chapter-title-sanskrit">{SITE_TITLE_SANSKRIT}</div>',
        '',
    ]
    parts += pending_banner(lang, [v for items in chapters.values() for v in items])
    parts += [
        '<div class="site-note">',
        esc(cfg['site_note']),
        '</div>',
        '',
        '<div class="site-usage">',
        esc(ui['usage']),
        '</div>',
        '',
        '<div class="site-download">',
        f'<a href="../{esc(ui["pdf_filename"])}">{esc(ui["download"])}</a>',
        '</div>',
        '',
        '<div class="chapter-grid">',
    ]
    parts += cards
    parts += ['</div>', '']
    (OUT / lang / 'index.qmd').write_text('\n'.join(parts), encoding='utf-8')


def write_glossary(lang, glossaries, verses):
    glossary = glossaries[lang]
    cfg = LANGS[lang]
    ui = cfg['ui']
    forms = glossary.forms_used(translation_for(v, lang)[0] for v in verses)
    groups = defaultdict(list)
    for item in glossary.entries:
        groups[item['group'] or cfg['glossary_groups'][0]].append(item)

    ordered = list(cfg['glossary_groups'])
    ordered += sorted(g for g in groups if g not in ordered)

    parts = page_header(ui['glossary'], lang, sidebar_id=lang)
    parts += [GENERATED_FILE_NOTE, '']
    parts += glossary_banner(lang, glossaries)
    parts += [
        '<div class="site-note glossary-note">',
        esc(cfg['glossary_note']),
        '</div>',
        '',
    ]
    for group in ordered:
        if group not in groups:
            continue
        parts += [f'## {group}', '', '<div class="glossary-list">']
        for item in sorted(groups[group], key=lambda x: x['term'].casefold()):
            variants = ', '.join(
                sorted(set(forms.get(item['id'], ())), key=lambda v: v.casefold())
            )
            parts += [
                f'<section class="glossary-entry" id="glossary-{esc(item["id"])}">',
                f'<h3>{esc(item["term"])}</h3>',
                f'<p>{esc(item["definition"])}</p>',
            ]
            # An entry whose term never surfaces in this translation (an epithet
            # rendered as plain wording, say) simply has no forms to list.
            if variants:
                parts.append(
                    f'<p class="glossary-variants">{esc(ui["glossary_forms"])} {esc(variants)}</p>'
                )
            parts.append('</section>')
        parts += ['</div>', '']
    (OUT / lang / 'glossary.qmd').write_text('\n'.join(parts), encoding='utf-8')


def write_chapters(lang, chapters, glossary):
    ui = LANGS[lang]['ui']
    chapter_dir = OUT / lang / 'chapters'
    chapter_dir.mkdir(parents=True, exist_ok=True)
    for old in chapter_dir.glob('chapter-*.qmd'):
        old.unlink()

    sorted_ch = sorted(chapters)
    for pos, ch in enumerate(sorted_ch):
        prev_ch = sorted_ch[pos - 1] if pos > 0 else None
        next_ch = sorted_ch[pos + 1] if pos < len(sorted_ch) - 1 else None
        items = chapters[ch]
        title_roman = CHAPTER_TITLES_ROMAN.get(ch, f'{ui["chapter"]} {ch}')
        title_sanskrit = items[0]['chapter_title_sanskrit'] if items else ''
        prev_link = (
            f'<a href="chapter-{prev_ch:02d}.html">{esc(ui["previous_chapter"])}</a>'
            if prev_ch else '<span></span>'
        )
        next_link = (
            f'<a href="chapter-{next_ch:02d}.html">{esc(ui["next_chapter"])}</a>'
            if next_ch else '<span></span>'
        )
        parts = page_header(f'{ui["chapter"]} {ch} - {title_roman}', lang, sidebar_id=lang)
        parts += [GENERATED_FILE_NOTE, '']
        if title_sanskrit:
            parts += [
                f'<div class="chapter-title-sanskrit">{html_lines(title_sanskrit)}</div>',
                '',
            ]
        parts += pending_banner(lang, items)
        parts += [
            f'<nav class="chapter-nav">{prev_link}{next_link}</nav>',
            '',
            '<div class="verse-list">',
        ]
        for v in items:
            ref = esc(v['reference'])
            anchor = f'v-{v["chapter"]}-{v["verse"]}'
            parts += [
                f'<article class="verse" id="{anchor}" data-reference="{ref}">',
                '<header class="verse-head">',
                f'<a class="ref" href="#{anchor}">{ref}</a>',
                f'<span class="speaker">{esc(v["speaker"])}</span>' if v['speaker'] else '',
                '</header>',
            ]
            text, fell_back = translation_for(v, lang)
            # Marking the paragraph with its real language keeps screen readers
            # and CSS honest when an untranslated verse shows the English text.
            fallback_attr = f' lang="{DEFAULT_LANG}"' if fell_back else ''
            parts += [
                f'<p class="translation translation-{lang}" data-lang="{lang}"{fallback_attr}>'
                f'{html_lines(text, glossary=glossary)}</p>',
            ]
            if v['sanskrit_sloka']:
                parts += [
                    '<details class="sanskrit">',
                    f'<summary>{esc(ui["sanskrit"])}</summary>',
                    f'<div class="sanskrit-text">{html_lines(v["sanskrit_sloka"])}</div>',
                    '</details>',
                ]
            parts += ['</article>', '']
        parts += ['</div>', '']
        (chapter_dir / f'chapter-{ch:02d}.qmd').write_text('\n'.join(parts), encoding='utf-8')


# ---------------------------------------------------------------------------
# Assets and data side-cars
# ---------------------------------------------------------------------------


def copy_flags():
    FLAG_OUT_DIR.mkdir(parents=True, exist_ok=True)
    for lang in LANG_ORDER:
        name = LANGS[lang]['flag']
        src = FLAG_SRC_DIR / name
        if not src.exists():
            raise FileNotFoundError(f'Missing flag image: {src}')
        shutil.copyfile(src, FLAG_OUT_DIR / name)


def write_chapter_titles_json(chapters):
    data = [
        {
            'chapter': ch,
            'title_roman': CHAPTER_TITLES_ROMAN.get(ch, f'Chapter {ch}'),
            'title_sanskrit': chapters[ch][0]['chapter_title_sanskrit'] if chapters[ch] else '',
            'verses': len(chapters[ch]),
        }
        for ch in sorted(chapters)
    ]
    (OUT / 'data' / 'chapter_titles.json').write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8'
    )


def remove_legacy_pages():
    """Earlier versions kept the English pages at the project root."""
    legacy = [OUT / 'glossary.qmd']
    legacy_dir = OUT / 'chapters'
    if legacy_dir.is_dir():
        for old in legacy_dir.glob('chapter-*.qmd'):
            old.unlink()
        if not any(legacy_dir.iterdir()):
            legacy_dir.rmdir()
    for path in legacy:
        if path.exists():
            path.unlink()


def report_progress(verses, chapters, lang_cols):
    """Print how far each translation has got, so the build doubles as a status
    check while the spreadsheet is being filled in."""
    total = len(verses)
    for lang in LANG_ORDER:
        if lang == DEFAULT_LANG:
            continue
        col = lang_cols.get(lang)
        source = f'column "{col}"' if col else 'no column in the spreadsheet yet'
        done = sum(1 for v in verses if not translation_for(v, lang)[1])
        pct = (100 * done / total) if total else 0
        print(f'  {lang}: {done}/{total} verses translated ({pct:.0f}%), from {source}')
        if done == total:
            continue
        complete, partial = [], []
        for ch in sorted(chapters):
            items = chapters[ch]
            n = sum(1 for v in items if not translation_for(v, lang)[1])
            if n == len(items):
                complete.append(ch)
            elif n:
                partial.append(f'{ch} ({n}/{len(items)})')
        if complete:
            print(f'      complete chapters: {", ".join(map(str, complete))}')
        if partial:
            print(f'      in progress: {", ".join(partial)}')
        if not complete and not partial:
            print('      nothing yet: the Italian pages show the English text')


# ---------------------------------------------------------------------------


def main():
    verses, lang_cols = read_verses(SRC_XLSX)
    chapters = defaultdict(list)
    for v in verses:
        chapters[v['chapter']].append(v)
    for ch in chapters:
        chapters[ch].sort(key=lambda x: x['verse'])

    (OUT / 'data').mkdir(exist_ok=True)
    with open(OUT / 'data' / 'verses.json', 'w', encoding='utf-8') as f:
        json.dump(verses, f, ensure_ascii=False, indent=2)

    copy_flags()
    remove_legacy_pages()
    write_quarto_yml(chapters)
    write_lang_config(chapters)
    write_root_index()
    write_chapter_titles_json(chapters)

    # Every glossary is loaded up front: a language needs the default one to
    # tell whether its own is still a verbatim copy.
    glossaries = {lang: Glossary(LANGS[lang]['glossary_csv']) for lang in LANG_ORDER}

    for lang in LANG_ORDER:
        glossary = glossaries[lang]
        (OUT / lang).mkdir(exist_ok=True)
        with open(OUT / 'data' / f'glossary.{lang}.json', 'w', encoding='utf-8') as f:
            json.dump(glossary.payload(), f, ensure_ascii=False, indent=2)
        if lang == DEFAULT_LANG:
            # Kept for tools/2_generate_pdf_book.py, which builds the English PDF.
            with open(OUT / 'data' / 'glossary.json', 'w', encoding='utf-8') as f:
                json.dump(glossary.payload(), f, ensure_ascii=False, indent=2)
        write_index(lang, chapters)
        write_glossary(lang, glossaries, verses)
        write_chapters(lang, chapters, glossary)

    print(f'Generated {len(verses)} verses in {len(chapters)} chapters '
          f'for {len(LANG_ORDER)} languages: {", ".join(LANG_ORDER)}.')
    report_progress(verses, chapters, lang_cols)


if __name__ == '__main__':
    main()
