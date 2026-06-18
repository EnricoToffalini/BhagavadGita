"""Patch an existing Bhagavad Gita Quarto project to show chapter titles.

Run from the project root, after the usual page generation:
    python tools/generate_site.py
    python add_chapter_titles.py

The script reads data/bhagavadgita_ai_refined.xlsx and updates:
- _quarto.yml
- index.qmd
- chapters/chapter-01.qmd ... chapters/chapter-18.qmd
- data/chapter_titles.json

It does not change verse text, Sanskrit verse blocks, or glossary tooltips.
"""

from pathlib import Path
from openpyxl import load_workbook
import html
import json
import re

ROOT = Path.cwd()
XLSX = ROOT / "data" / "bhagavadgita_ai_refined.xlsx"

# Conventional romanized chapter titles, aligned with the Sanskrit titles in the spreadsheet.
# Some editions use close variants, for example Karma-Sannyasa Yoga for Chapter 5 or Dhyana Yoga for Chapter 6.
CHAPTER_TITLES_ROMAN = {
    1: "Arjuna Vishada Yoga",
    2: "Sankhya Yoga",
    3: "Karma Yoga",
    4: "Jnana-Karma-Sannyasa Yoga",
    5: "Sannyasa Yoga",
    6: "Atma-Samyama Yoga",
    7: "Jnana-Vijnana Yoga",
    8: "Akshara-Brahma Yoga",
    9: "Raja-Vidya Raja-Guhya Yoga",
    10: "Vibhuti Yoga",
    11: "Vishvarupa-Darshana Yoga",
    12: "Bhakti Yoga",
    13: "Kshetra-Kshetrajna Vibhaga Yoga",
    14: "Guna-Traya Vibhaga Yoga",
    15: "Purushottama Yoga",
    16: "Daivasura-Sampad Vibhaga Yoga",
    17: "Shraddha-Traya Vibhaga Yoga",
    18: "Moksha-Sannyasa Yoga",
}


def esc(x):
    return html.escape(str(x or "").strip(), quote=True)


def read_chapter_data():
    if not XLSX.exists():
        raise FileNotFoundError(f"Could not find {XLSX}. Run this script from the Quarto project root.")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["chapter", "chapter_title_sanskrit", "verse"]
    missing = [col for col in required if col not in idx]
    if missing:
        raise ValueError(f"Missing required columns in {XLSX}: {missing}")

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ch = row[idx["chapter"]]
        if ch is None:
            continue
        ch = int(ch)
        data.setdefault(ch, {
            "chapter": ch,
            "title_roman": CHAPTER_TITLES_ROMAN.get(ch, f"Chapter {ch}"),
            "title_sanskrit": str(row[idx["chapter_title_sanskrit"]] or "").strip(),
            "verses": 0,
        })
        data[ch]["verses"] += 1
    return dict(sorted(data.items()))


def write_json(data):
    out = ROOT / "data" / "chapter_titles.json"
    out.write_text(json.dumps(list(data.values()), ensure_ascii=False, indent=2), encoding="utf-8")


def remove_div_block(text, marker):
    start = text.find(marker)
    if start == -1:
        return text

    depth = 0
    pattern = re.compile(r'</?div\b[^>]*>', flags=re.I)
    for match in pattern.finditer(text, start):
        tag = match.group(0).lower()
        if tag.startswith('<div'):
            depth += 1
        else:
            depth -= 1
            if depth == 0:
                end = match.end()
                while end < len(text) and text[end] in ' \t\r\n':
                    end += 1
                return text[:start] + text[end:]
    return text


def patch_quarto_yml(data):
    path = ROOT / "_quarto.yml"
    if not path.exists():
        return
    text = path.read_text(encoding="utf-8")
    for ch, info in data.items():
        title = info["title_roman"]
        # Matches both navbar and sidebar entries generated as text: "Chapter N".
        text = re.sub(
            rf'(\s*- text: )"Chapter {ch}"(\n\s+href: chapters/chapter-{ch:02d}\.qmd)',
            rf'\1"{ch} · {title}"\2',
            text,
        )
    path.write_text(text, encoding="utf-8")


def patch_index(data):
    path = ROOT / "index.qmd"
    if not path.exists():
        return
    text = path.read_text(encoding="utf-8")
    for ch, info in data.items():
        title = esc(info["title_roman"])
        # Remove an old inserted title, then insert the current one after the chapter number.
        text = re.sub(
            rf'(<span class="chapter-number">Chapter {ch}</span>)<span class="chapter-title">.*?</span>',
            rf'\1',
            text,
            flags=re.S,
        )
        text = text.replace(
            f'<span class="chapter-number">Chapter {ch}</span>',
            f'<span class="chapter-number">Chapter {ch}</span><span class="chapter-title">{title}</span>',
        )
    path.write_text(text, encoding="utf-8")


def patch_chapter_pages(data):
    chapter_dir = ROOT / "chapters"
    for ch, info in data.items():
        path = chapter_dir / f"chapter-{ch:02d}.qmd"
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8")
        roman = esc(info["title_roman"])
        sanskrit = esc(info["title_sanskrit"])

        # Keep the H1 self-contained and remove visible metadata duplication.
        text = re.sub(
            rf'^(title: )"Chapter {ch}"\s*$',
            rf'\1"Chapter {ch} - {roman}"',
            text,
            count=1,
            flags=re.M,
        )
        text = re.sub(
            rf'^(title: )"Chapter {ch} - .*?"\s*$',
            rf'\1"Chapter {ch} - {roman}"',
            text,
            count=1,
            flags=re.M,
        )
        text = re.sub(r'^subtitle: ".*?"\n?', '', text, count=1, flags=re.M)
        text = re.sub(r'^description: ".*?"\n?', '', text, count=1, flags=re.M)

        # Remove the old chapter metadata block entirely.
        text = remove_div_block(text, '<div class="chapter-meta">')

        # Remove any previously inserted top-level Sanskrit title line before re-adding it.
        text = remove_div_block(text, '<div class="chapter-title-sanskrit-inline"')
        text = remove_div_block(text, '<div class="chapter-title-sanskrit">')

        if sanskrit:
            sanskrit_block = f'<div class="chapter-title-sanskrit">{sanskrit}</div>'
            text = re.sub(
                r'^(---\n.*?\n---\n\n)',
                rf'\1{sanskrit_block}\n\n',
                text,
                count=1,
                flags=re.S,
            )
        path.write_text(text, encoding="utf-8")


def main():
    data = read_chapter_data()
    write_json(data)
    patch_quarto_yml(data)
    patch_index(data)
    patch_chapter_pages(data)
    print(f"Patched chapter titles for {len(data)} chapters.")
    print("Created data/chapter_titles.json.")


if __name__ == "__main__":
    main()
